"""
Excel order parser for embroidery combo workflow.
Reads an order Excel, groups names by (machine_program, com_no),
expands quantities, and splits into combo files of max 20 slots.

Supports auto-detection of column positions from headers.
"""

from collections import defaultdict
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

from copy import copy

from openpyxl import load_workbook

# Default column indices (A=0, F=5, G=6, H=7, O=14, P=15)
DEFAULT_COLUMN_MAP = {
    "program": 0,
    "name_line1": 5,
    "name_line2": 6,
    "quantity": 7,
    "com_no": 14,
    "machine_program": 15,
}

# Header patterns for auto-detection (lowercase). Order matters — first match wins.
HEADER_PATTERNS = {
    "program": ["program", "prog", "prg", "file number", "dst"],
    "name_line1": ["row 1", "name 1", "first name", "embroidery name", "name"],
    "name_line2": ["row 2", "name 2", "title", "subtitle", "organisation", "organization"],
    "quantity": ["quantity", "qty", "amount", "count", "pcs"],
    "com_no": ["com no", "combo no", "combo number", "combo", "com"],
    "machine_program": ["machine program", "machine prog", "machine", "m/a", "ma"],
}

# Header patterns for auto-assign columns (size, fabric, frame, embroidery)
ASSIGN_HEADER_PATTERNS = {
    "size": ["size", "sz"],
    "fabric_colour": ["fabric color", "fabric colour", "fabric col", "fabric"],
    "frame_colour": ["frame color", "frame colour", "frame col", "frame"],
    "embroidery_colour": ["name/embroidery", "embroidery color", "embroidery colour", "embroidery col", "embroidery"],
}

# Default column indices for auto-assign (I=8, J=9, K=10, M=12)
DEFAULT_ASSIGN_COLUMN_MAP = {
    "size": 12,
    "fabric_colour": 8,
    "frame_colour": 9,
    "embroidery_colour": 10,
}

# Fields where "M" alone is a valid header (special case — too short for fuzzy match)
M_HEADER_FIELD = "machine_program"


@dataclass
class NameEntry:
    program: int
    name_line1: str
    name_line2: str
    quantity: int
    com_no: str
    machine_program: str


@dataclass
class ComboGroup:
    machine_program: str
    com_no: str
    entries: List[NameEntry] = field(default_factory=list)

    @property
    def group_key(self) -> Tuple[str, str]:
        return (self.machine_program, self.com_no)

    @property
    def total_slots(self) -> int:
        return sum(e.quantity for e in self.entries)


@dataclass
class ComboFile:
    machine_program: str
    com_no: str
    part_number: int
    total_parts: int
    slots: List[NameEntry] = field(default_factory=list)
    head_mode: str = ""  # "" = legacy, "1-HEAD" or "2-HEAD" = optimized

    @property
    def left_column(self) -> List[NameEntry]:
        return self.slots[:10]

    @property
    def right_column(self) -> List[NameEntry]:
        return self.slots[10:]

    @property
    def filename(self) -> str:
        if self.head_mode == "2-HEAD":
            # Show original qty so workers know how many per name
            # Even files are grouped by qty, so all slots have the same quantity
            qty_label = f"_Qty{self.slots[0].quantity}" if self.slots else ""
            tag = f"_EVEN{qty_label}"
        elif self.head_mode == "1-HEAD":
            tag = "_ODD"
        else:
            tag = ""
        return f"{self.machine_program}_Com{self.com_no}{tag}_{self.part_number}of{self.total_parts}.dst"


@dataclass
class ParseResult:
    entries: List[NameEntry] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


@dataclass
class DetectResult:
    headers: List[str]
    preview_rows: List[List]
    detected_mapping: Dict[str, int]
    confidence: str  # "high" or "low"


def detect_columns(path: str, preview_count: int = 5) -> DetectResult:
    """Read Excel headers and auto-detect column mapping.

    Returns headers, sample rows, detected mapping, and confidence level.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return DetectResult([], [], dict(DEFAULT_COLUMN_MAP), "low")

    # Row 1 = headers
    raw_headers = list(rows[0]) if rows else []
    headers = [str(h or "").strip() for h in raw_headers]
    headers_lower = [h.lower() for h in headers]

    # Sample data rows (skip header)
    preview_rows = []
    for row in rows[1:preview_count + 1]:
        preview_rows.append([_cell_to_json(c) for c in row])

    # Auto-detect mapping
    mapping: Dict[str, int] = {}
    used_indices: set = set()

    # Special case: header exactly "M" maps to machine_program
    for i, h in enumerate(headers):
        if h.strip() == "M" and M_HEADER_FIELD not in mapping:
            mapping[M_HEADER_FIELD] = i
            used_indices.add(i)
            break

    # Match each field by header patterns
    for field_name, patterns in HEADER_PATTERNS.items():
        if field_name in mapping:
            continue  # already matched (e.g., machine_program via "M")
        for pattern in patterns:
            for i, h in enumerate(headers_lower):
                if i in used_indices:
                    continue
                if pattern == h or (len(pattern) > 2 and pattern in h):
                    # For "program", skip if this looks like a duplicate (column N often duplicates A)
                    if field_name == "program" and i > 0 and "program" in mapping:
                        continue
                    # "com" should not match "commission" — only "com", "com no", etc.
                    if field_name == "com_no" and "commission" in h:
                        continue
                    mapping[field_name] = i
                    used_indices.add(i)
                    break
            if field_name in mapping:
                break

    # Fill missing fields from defaults
    all_fields = list(DEFAULT_COLUMN_MAP.keys())
    matched = len(mapping)
    for f in all_fields:
        if f not in mapping:
            mapping[f] = DEFAULT_COLUMN_MAP[f]

    confidence = "high" if matched >= 5 else ("low" if matched < 3 else "medium")

    return DetectResult(
        headers=headers,
        preview_rows=preview_rows,
        detected_mapping=mapping,
        confidence=confidence,
    )


def _cell_to_json(val):
    """Convert a cell value to a JSON-safe type."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and val == int(val):
            return int(val)
        return val
    return str(val)


def parse_excel(path: str, column_map: Optional[Dict[str, int]] = None) -> ParseResult:
    """Read Excel and extract entries using column mapping.

    If column_map is not provided, uses DEFAULT_COLUMN_MAP.
    """
    cmap = column_map or DEFAULT_COLUMN_MAP
    idx_program = cmap["program"]
    idx_name1 = cmap["name_line1"]
    idx_name2 = cmap["name_line2"]
    idx_qty = cmap["quantity"]
    idx_com = cmap["com_no"]
    idx_m = cmap["machine_program"]
    max_idx = max(idx_program, idx_name1, idx_name2, idx_qty, idx_com, idx_m)

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    result = ParseResult()

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) <= max_idx:
            continue

        program = row[idx_program]
        name1 = row[idx_name1]
        name2 = row[idx_name2]
        qty = row[idx_qty]
        com_no = row[idx_com]
        m_val = row[idx_m]

        if program is None or m_val is None:
            if program is not None or name1 is not None:
                result.warnings.append(f"Row {row_num}: missing program or M value, skipped")
            continue

        try:
            program = int(program)
        except (ValueError, TypeError):
            result.warnings.append(f"Row {row_num}: program '{program}' is not a number, skipped")
            continue

        # Robust quantity parsing: handle text like "50 pcs", cap at 1000
        if qty is None:
            qty = 1
        else:
            try:
                qty = int(float(qty))  # handles "50.0" and plain numbers
            except (ValueError, TypeError):
                # Try extracting leading digits from text like "50 pcs"
                import re as _re
                m = _re.match(r'(\d+)', str(qty).strip())
                if m:
                    result.warnings.append(f"Row {row_num}: qty '{qty}' → using {m.group(1)}")
                    qty = int(m.group(1))
                else:
                    result.warnings.append(f"Row {row_num}: qty '{qty}' is not a number, using 1")
                    qty = 1
            if qty < 1:
                result.warnings.append(f"Row {row_num}: qty={qty} treated as 1")
                qty = 1
            elif qty > 1000:
                result.warnings.append(f"Row {row_num}: qty={qty} capped at 1000")
                qty = 1000

        result.entries.append(NameEntry(
            program=program,
            name_line1=str(name1 or ""),
            name_line2=str(name2 or ""),
            quantity=qty,
            com_no=str(int(com_no)) if com_no is not None and isinstance(com_no, (int, float)) else str(com_no or ""),
            machine_program=str(m_val),
        ))

    wb.close()
    return result


def group_entries(entries: List[NameEntry]) -> List[ComboGroup]:
    """Group by (machine_program, com_no), sort entries within each group by program number."""
    groups_dict = defaultdict(list)
    for entry in entries:
        key = (entry.machine_program, entry.com_no)
        groups_dict[key].append(entry)

    groups = []
    for (m, com), ents in sorted(groups_dict.items()):
        ents.sort(key=lambda e: e.program)
        groups.append(ComboGroup(machine_program=m, com_no=com, entries=ents))
    return groups


def expand_and_split(group: ComboGroup, max_slots: int = 20) -> List[ComboFile]:
    """
    Expand quantities and split into combo files of max_slots each.

    Sorts by (quantity DESC, program ASC) to cluster same-qty entries.
    Each entry is repeated by its quantity in the expanded slot list.
    Splits at max_slots boundaries.
    """
    sorted_entries = sorted(group.entries, key=lambda e: (-e.quantity, e.program))

    expanded = []
    for entry in sorted_entries:
        for _ in range(entry.quantity):
            expanded.append(entry)

    if not expanded:
        return []

    chunks = []
    for i in range(0, len(expanded), max_slots):
        chunks.append(expanded[i:i + max_slots])

    total_parts = len(chunks)
    combo_files = []
    for i, chunk in enumerate(chunks, 1):
        combo_files.append(ComboFile(
            machine_program=group.machine_program,
            com_no=group.com_no,
            part_number=i,
            total_parts=total_parts,
            slots=chunk,
        ))
    return combo_files


def expand_and_split_with_heads(group: ComboGroup, max_slots: int = 20) -> List[ComboFile]:
    """
    Expand quantities with 2-HEAD optimization.

    - Even-qty entries: slots = qty / 2 (both heads run same file)
    - Odd-qty entries: slots = qty (single head)
    - Even entries are further grouped by quantity so each file has one qty label.
    - Even and odd entries get SEPARATE combo files within the same group.
    """
    even_entries = [e for e in group.entries if e.quantity >= 2 and e.quantity % 2 == 0]
    odd_entries = [e for e in group.entries if e.quantity % 2 != 0 or e.quantity < 2]

    combo_files = []

    # --- Process EVEN entries, grouped by quantity (2-HEAD mode, half slots) ---
    if even_entries:
        # Group by quantity so each file only has names of the same qty
        qty_groups: Dict[int, List[NameEntry]] = defaultdict(list)
        for e in even_entries:
            qty_groups[e.quantity].append(e)

        for qty in sorted(qty_groups.keys()):
            entries_for_qty = sorted(qty_groups[qty], key=lambda e: e.program)
            expanded = []
            for entry in entries_for_qty:
                for _ in range(entry.quantity // 2):
                    expanded.append(entry)

            chunks = [expanded[i:i + max_slots] for i in range(0, len(expanded), max_slots)]
            total_parts = len(chunks)
            for i, chunk in enumerate(chunks, 1):
                combo_files.append(ComboFile(
                    machine_program=group.machine_program,
                    com_no=group.com_no,
                    part_number=i,
                    total_parts=total_parts,
                    slots=chunk,
                    head_mode="2-HEAD",
                ))

    # --- Process ODD entries (1-HEAD mode, full slots) ---
    if odd_entries:
        sorted_odd = sorted(odd_entries, key=lambda e: (-e.quantity, e.program))
        expanded_odd = []
        for entry in sorted_odd:
            for _ in range(entry.quantity):
                expanded_odd.append(entry)

        chunks = [expanded_odd[i:i + max_slots] for i in range(0, len(expanded_odd), max_slots)]
        total_parts = len(chunks)
        for i, chunk in enumerate(chunks, 1):
            combo_files.append(ComboFile(
                machine_program=group.machine_program,
                com_no=group.com_no,
                part_number=i,
                total_parts=total_parts,
                slots=chunk,
                head_mode="1-HEAD",
            ))

    return combo_files


def generate_all_combos(entries: List[NameEntry], max_slots: int = 20, optimize_heads: bool = False) -> List[ComboFile]:
    """Full pipeline: group -> expand -> split. Returns all combo files."""
    groups = group_entries(entries)
    all_combos = []
    for group in groups:
        if optimize_heads:
            all_combos.extend(expand_and_split_with_heads(group, max_slots))
        else:
            all_combos.extend(expand_and_split(group, max_slots))
    return all_combos


# ---------------------------------------------------------------------------
# Auto-assign MA & COM
# ---------------------------------------------------------------------------

@dataclass
class AutoAssignResult:
    """Result of auto-assigning MA and COM numbers."""
    headers: List[str]
    preview_rows: List[List]
    detected_mapping: Dict[str, int]  # size, fabric_colour, frame_colour, embroidery_colour
    confidence: str
    assignments: List[Dict]  # per-row: {row_num, size, fabric, frame, embroidery, ma, com}
    ma_summary: List[Dict]   # {ma, size, count}
    com_summary: List[Dict]  # {ma, com, fabric, frame, embroidery, count}
    warnings: List[str] = field(default_factory=list)


def detect_assign_columns(path: str, preview_count: int = 5) -> AutoAssignResult:
    """Read Excel headers and auto-detect columns for MA/COM assignment.

    Detects: size, fabric_colour, frame_colour, embroidery_colour.
    Returns headers, preview rows, detected mapping, and empty assignments
    (call auto_assign_ma_com to populate assignments).
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return AutoAssignResult([], [], dict(DEFAULT_ASSIGN_COLUMN_MAP), "low", [], [], [])

    raw_headers = list(rows[0])
    headers = [str(h or "").strip() for h in raw_headers]
    headers_lower = [h.lower() for h in headers]

    preview_rows = []
    for row in rows[1:preview_count + 1]:
        preview_rows.append([_cell_to_json(c) for c in row])

    # Auto-detect mapping
    mapping: Dict[str, int] = {}
    used_indices: set = set()

    for field_name, patterns in ASSIGN_HEADER_PATTERNS.items():
        for pattern in patterns:
            for i, h in enumerate(headers_lower):
                if i in used_indices:
                    continue
                if pattern == h or (len(pattern) > 2 and pattern in h):
                    mapping[field_name] = i
                    used_indices.add(i)
                    break
            if field_name in mapping:
                break

    matched = len(mapping)
    for f in DEFAULT_ASSIGN_COLUMN_MAP:
        if f not in mapping:
            mapping[f] = DEFAULT_ASSIGN_COLUMN_MAP[f]

    confidence = "high" if matched >= 3 else ("low" if matched < 2 else "medium")

    return AutoAssignResult(
        headers=headers,
        preview_rows=preview_rows,
        detected_mapping=mapping,
        confidence=confidence,
        assignments=[],
        ma_summary=[],
        com_summary=[],
    )


def auto_assign_ma_com(
    path: str,
    column_map: Optional[Dict[str, int]] = None,
    ma_lookup: Optional[Dict[str, str]] = None,
    com_lookup: Optional[Dict[str, Dict[tuple, int]]] = None,
    max_com_per_ma: Optional[Dict[str, int]] = None,
) -> AutoAssignResult:
    """Auto-assign MA and COM numbers based on size and colour columns.

    MA: looked up from ma_lookup (normalized_size → real MA number) if provided,
        otherwise flags as NEW with a placeholder.
    COM: looked up from com_lookup (ma → {colour_key → com_number}) if provided,
         otherwise assigns next available COM and flags as NEW.
    """
    cmap = column_map or DEFAULT_ASSIGN_COLUMN_MAP
    idx_size = cmap["size"]
    idx_fabric = cmap["fabric_colour"]
    idx_frame = cmap["frame_colour"]
    idx_embroidery = cmap["embroidery_colour"]
    max_idx = max(idx_size, idx_fabric, idx_frame, idx_embroidery)

    # Try to detect quantity column for 2-HEAD optimization
    idx_quantity = None

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return AutoAssignResult([], [], cmap, "low", [], [], [])

    raw_headers = list(rows[0])
    headers = [str(h or "").strip() for h in raw_headers]
    headers_lower = [h.lower() for h in headers]
    preview_rows = []
    for row in rows[1:6]:
        preview_rows.append([_cell_to_json(c) for c in row])

    # Auto-detect quantity column from headers
    qty_patterns = ["quantity", "qty", "amount", "count", "pcs"]
    for i, h in enumerate(headers_lower):
        for pat in qty_patterns:
            if pat in h:
                idx_quantity = i
                break
        if idx_quantity is not None:
            break

    # --- Assignment algorithm ---
    size_to_ma: Dict[str, str] = {}        # normalized_size → MA number
    size_display: Dict[str, str] = {}      # normalized_size → first raw size seen
    ma_is_new: Dict[str, bool] = {}        # normalized_size → True if MA is new/unknown
    new_ma_counter = 1
    # Track COM: ma → {colour_key → com_number}
    com_tracker: Dict[str, Dict[Tuple, int]] = {}
    com_is_new_tracker: Dict[str, Dict[Tuple, bool]] = {}  # ma → {colour_key → is_new}
    # Track next COM number per MA (continue from reference max)
    next_com: Dict[str, int] = {}
    if max_com_per_ma:
        for ma, max_c in max_com_per_ma.items():
            next_com[ma] = max_c + 1

    # Pre-populate com_tracker from reference so existing COMs are reused
    if com_lookup:
        for ma, colour_dict in com_lookup.items():
            com_tracker[ma] = {}
            com_is_new_tracker[ma] = {}
            for colour_key, com_no in colour_dict.items():
                com_tracker[ma][colour_key] = com_no
                com_is_new_tracker[ma][colour_key] = False

    assignments = []
    warnings = []

    for row_num, row in enumerate(rows[1:], start=2):
        if len(row) <= max_idx:
            continue

        size_raw = str(row[idx_size] or "").strip()
        fabric_raw = str(row[idx_fabric] or "").strip()
        frame_raw = str(row[idx_frame] or "").strip()
        embroidery_raw = str(row[idx_embroidery] or "").strip()

        if not size_raw:
            warnings.append(f"Row {row_num}: missing size value, skipped")
            continue

        # Normalize for grouping (collapse whitespace, case-insensitive)
        import re as _re
        size_val = _re.sub(r'\s+', '', size_raw).lower()  # "110 x 35" → "110x35"
        fabric_val = fabric_raw.strip().title()
        frame_val = frame_raw.strip().title()
        embroidery_val = embroidery_raw.strip().title()

        # Assign MA
        is_new_ma = False
        if size_val not in size_to_ma:
            if ma_lookup and size_val in ma_lookup:
                size_to_ma[size_val] = ma_lookup[size_val]
            else:
                # New size — placeholder MA, flagged
                placeholder = f"NEW-MA{new_ma_counter}"
                size_to_ma[size_val] = placeholder
                new_ma_counter += 1
                is_new_ma = True
                warnings.append(f"⚠ Row {row_num}: size '{size_raw}' not in reference — assigned placeholder {placeholder}")
            size_display[size_val] = size_raw
            ma_is_new[size_val] = is_new_ma
        ma = size_to_ma[size_val]

        # Assign COM
        if ma not in com_tracker:
            com_tracker[ma] = {}
            com_is_new_tracker[ma] = {}
        colour_key = (fabric_val, frame_val, embroidery_val)

        is_new_com = False
        if colour_key not in com_tracker[ma]:
            # Not in reference — assign next available COM
            if ma not in next_com:
                # No existing COMs for this MA, start at 1
                next_com[ma] = 1
            com_tracker[ma][colour_key] = next_com[ma]
            next_com[ma] += 1
            is_new_com = True
            if not ma_is_new.get(size_val, False):
                # Only warn about new COM if the MA itself isn't new (avoid double-warning)
                warnings.append(
                    f"⚠ Row {row_num}: new color combo for {ma} — "
                    f"fabric={fabric_val}, frame={frame_val}, embroidery={embroidery_val} "
                    f"→ assigned COM {com_tracker[ma][colour_key]}"
                )
            com_is_new_tracker[ma][colour_key] = True
        else:
            is_new_com = com_is_new_tracker[ma].get(colour_key, False)

        com = com_tracker[ma][colour_key]

        # Read quantity if available
        qty = 1
        if idx_quantity is not None and idx_quantity < len(row):
            qty_raw = row[idx_quantity]
            if qty_raw is not None:
                try:
                    qty = int(float(str(qty_raw).split()[0]))
                except (ValueError, IndexError):
                    qty = 1
            if qty < 1:
                qty = 1

        assignments.append({
            "row_num": row_num,
            "size": size_raw,
            "fabric_colour": fabric_val,
            "frame_colour": frame_val,
            "embroidery_colour": embroidery_val,
            "assigned_ma": ma,
            "assigned_com": com,
            "is_new_ma": ma_is_new.get(size_val, False),
            "is_new_com": is_new_com,
            "quantity": qty,
        })

    # Build summaries
    ma_counts: Dict[str, int] = defaultdict(int)
    for a in assignments:
        ma_counts[a["assigned_ma"]] += 1

    ma_summary = []
    for size_val, ma in size_to_ma.items():
        ma_summary.append({
            "ma": ma,
            "size": size_display.get(size_val, size_val),
            "count": ma_counts[ma],
            "is_new": ma_is_new.get(size_val, False),
        })

    com_summary = []
    for ma, colour_dict in com_tracker.items():
        for (fabric, frame, embroidery), com_no in colour_dict.items():
            count = sum(1 for a in assignments if a["assigned_ma"] == ma and a["assigned_com"] == com_no)
            if count == 0:
                continue  # Skip reference COMs that aren't used in this order
            com_summary.append({
                "ma": ma,
                "com": com_no,
                "fabric_colour": fabric,
                "frame_colour": frame,
                "embroidery_colour": embroidery,
                "count": count,
                "is_new": com_is_new_tracker.get(ma, {}).get((fabric, frame, embroidery), False),
            })

    return AutoAssignResult(
        headers=headers,
        preview_rows=preview_rows,
        detected_mapping=cmap,
        confidence="high" if len(assignments) > 0 else "low",
        assignments=assignments,
        ma_summary=ma_summary,
        com_summary=com_summary,
        warnings=warnings,
    )


def export_assigned_excel(
    original_path: str,
    assignments: List[Dict],
    ma_summary: List[Dict] = None,
    com_summary: List[Dict] = None,
    com_col: int = None,
    ma_col: int = None,
    optimize_heads: bool = False,
) -> str:
    """Write a copy of the Excel with MA and COM columns, flagged rows, and a summary sheet.

    - Appends MA, COM, and Status columns to the order sheet
    - Highlights NEW MA/COM rows in yellow
    - Adds a 'Summary' sheet with full mapping breakdown
    Returns the path to the new file.
    """
    from openpyxl import load_workbook as _load_wb
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import os

    wb = _load_wb(original_path)
    ws = wb.active

    # --- Styles ---
    bold = Font(bold=True)
    bold_white = Font(bold=True, color="FFFFFF")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    light_yellow = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")
    header_fill = PatternFill(start_color="3B4985", end_color="3B4985", fill_type="solid")
    blue_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # --- Find last used column ---
    if ma_col is None or com_col is None:
        last_col = 0
        for cell in ws[1]:
            if cell.value is not None:
                last_col = cell.column
        ma_col_1 = last_col + 1
        com_col_1 = last_col + 2
        status_col_1 = last_col + 3
    else:
        ma_col_1 = ma_col + 1
        com_col_1 = com_col + 1
        status_col_1 = max(ma_col_1, com_col_1) + 1

    # Extra columns for 2-HEAD optimization
    mode_col_1 = status_col_1 + 1 if optimize_heads else None
    slots_col_1 = status_col_1 + 2 if optimize_heads else None

    # --- Write headers ---
    header_cols = [(ma_col_1, "MA"), (com_col_1, "COM"), (status_col_1, "Status")]
    if optimize_heads:
        header_cols += [(mode_col_1, "Mode"), (slots_col_1, "Slots")]
    for col, label in header_cols:
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = bold_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions[get_column_letter(ma_col_1)].width = 14
    ws.column_dimensions[get_column_letter(com_col_1)].width = 8
    ws.column_dimensions[get_column_letter(status_col_1)].width = 18
    if optimize_heads:
        ws.column_dimensions[get_column_letter(mode_col_1)].width = 10
        ws.column_dimensions[get_column_letter(slots_col_1)].width = 8

    # --- Build lookup: row_num -> assignment ---
    row_lookup = {a["row_num"]: a for a in assignments}

    new_count = 0
    for row_num, a in row_lookup.items():
        ma = a["assigned_ma"]
        com = a["assigned_com"]
        is_new_ma = a.get("is_new_ma", False)
        is_new_com = a.get("is_new_com", False)

        ma_cell = ws.cell(row=row_num, column=ma_col_1, value=str(ma))
        com_cell = ws.cell(row=row_num, column=com_col_1, value=com)
        status_cell = ws.cell(row=row_num, column=status_col_1)

        if is_new_ma:
            status_cell.value = "NEW MA + COM"
            for c in [ma_cell, com_cell, status_cell]:
                c.fill = yellow_fill
                c.font = Font(bold=True)
            new_count += 1
        elif is_new_com:
            status_cell.value = "NEW COM"
            for c in [ma_cell, com_cell, status_cell]:
                c.fill = light_yellow
                c.font = Font(bold=True)
            new_count += 1
        else:
            status_cell.value = "OK"
            status_cell.fill = green_fill

        # 2-HEAD optimization columns
        if optimize_heads and mode_col_1 and slots_col_1:
            qty = a.get("quantity", 1)
            if qty is None:
                qty = 1
            is_even = qty >= 2 and qty % 2 == 0
            mode_cell = ws.cell(row=row_num, column=mode_col_1, value="2-HEAD" if is_even else "1-HEAD")
            slots_cell = ws.cell(row=row_num, column=slots_col_1, value=qty // 2 if is_even else qty)
            mode_cell.alignment = Alignment(horizontal="center")
            slots_cell.alignment = Alignment(horizontal="center")
            if is_even:
                mode_cell.fill = blue_fill
                slots_cell.fill = blue_fill
                mode_cell.font = Font(bold=True)

    # =====================================================================
    # SUMMARY SHEET
    # =====================================================================
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    summary = wb.create_sheet("Summary")  # Append as last sheet (order data stays as active/first)

    def _write_header_row(sheet, row, headers, col_start=1):
        for i, h in enumerate(headers):
            cell = sheet.cell(row=row, column=col_start + i, value=h)
            cell.font = bold_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    def _write_cell(sheet, row, col, value, fill=None, font_style=None, align_center=False):
        cell = sheet.cell(row=row, column=col, value=value)
        cell.border = thin_border
        if fill:
            cell.fill = fill
        if font_style:
            cell.font = font_style
        if align_center:
            cell.alignment = Alignment(horizontal="center")
        return cell

    current_row = 1

    # --- Title ---
    title_cell = summary.cell(row=current_row, column=1, value="Micro Automation — Assignment Summary")
    title_cell.font = Font(bold=True, size=14, color="3B4985")
    current_row += 2

    # --- Overall stats ---
    total = len(assignments)
    matched = total - new_count
    summary.cell(row=current_row, column=1, value="Total order rows:").font = bold
    summary.cell(row=current_row, column=2, value=total)
    current_row += 1
    summary.cell(row=current_row, column=1, value="Matched from reference:").font = bold
    _write_cell(summary, current_row, 2, matched, fill=green_fill)
    current_row += 1
    summary.cell(row=current_row, column=1, value="New (flagged):").font = bold
    _write_cell(summary, current_row, 2, new_count, fill=yellow_fill if new_count > 0 else green_fill)
    current_row += 2

    # --- MA Summary Table ---
    summary.cell(row=current_row, column=1, value="MA Summary").font = Font(bold=True, size=12, color="3B4985")
    current_row += 1
    _write_header_row(summary, current_row, ["MA", "Size", "Order Rows", "Status"])
    current_row += 1

    if ma_summary:
        for m in sorted(ma_summary, key=lambda x: x["ma"]):
            is_new = m.get("is_new", False)
            fill = yellow_fill if is_new else None
            _write_cell(summary, current_row, 1, m["ma"], fill=fill, font_style=bold if is_new else None)
            _write_cell(summary, current_row, 2, m["size"], fill=fill)
            _write_cell(summary, current_row, 3, m["count"], fill=fill, align_center=True)
            status = "NEW — add to reference" if is_new else "OK"
            _write_cell(summary, current_row, 4, status, fill=yellow_fill if is_new else green_fill)
            current_row += 1

    current_row += 2

    # --- COM Breakdown Table ---
    summary.cell(row=current_row, column=1, value="COM Breakdown").font = Font(bold=True, size=12, color="3B4985")
    current_row += 1
    _write_header_row(summary, current_row, ["MA", "COM", "Fabric", "Frame", "Embroidery", "Order Rows", "Status"])
    current_row += 1

    if com_summary:
        for c in sorted(com_summary, key=lambda x: (x["ma"], x["com"])):
            is_new = c.get("is_new", False)
            fill = yellow_fill if is_new else None
            _write_cell(summary, current_row, 1, c["ma"], fill=fill, font_style=bold if is_new else None)
            _write_cell(summary, current_row, 2, c["com"], fill=fill, align_center=True)
            _write_cell(summary, current_row, 3, c.get("fabric_colour", ""), fill=fill)
            _write_cell(summary, current_row, 4, c.get("frame_colour", ""), fill=fill)
            _write_cell(summary, current_row, 5, c.get("embroidery_colour", ""), fill=fill)
            _write_cell(summary, current_row, 6, c.get("count", 0), fill=fill, align_center=True)
            status = "NEW — add to reference" if is_new else "OK"
            _write_cell(summary, current_row, 7, status, fill=yellow_fill if is_new else green_fill)
            current_row += 1

    current_row += 2

    # --- 2-HEAD Optimization Summary ---
    if optimize_heads:
        summary.cell(row=current_row, column=1, value="2-HEAD Optimization").font = Font(bold=True, size=12, color="3B4985")
        current_row += 1

        even_rows = [a for a in assignments if a.get("quantity", 1) >= 2 and a.get("quantity", 1) % 2 == 0]
        odd_rows = [a for a in assignments if a.get("quantity", 1) % 2 != 0 or a.get("quantity", 1) < 2]
        total_original = sum(a.get("quantity", 1) for a in assignments)
        total_optimized = sum(a.get("quantity", 1) // 2 for a in even_rows) + sum(a.get("quantity", 1) for a in odd_rows)
        slots_saved = total_original - total_optimized

        summary.cell(row=current_row, column=1, value="Even-qty rows (2-HEAD):").font = bold
        _write_cell(summary, current_row, 2, len(even_rows), fill=blue_fill)
        current_row += 1
        summary.cell(row=current_row, column=1, value="Odd-qty rows (1-HEAD):").font = bold
        _write_cell(summary, current_row, 2, len(odd_rows))
        current_row += 1
        summary.cell(row=current_row, column=1, value="Original total slots:").font = bold
        _write_cell(summary, current_row, 2, total_original)
        current_row += 1
        summary.cell(row=current_row, column=1, value="Optimized total slots:").font = bold
        _write_cell(summary, current_row, 2, total_optimized, fill=blue_fill, font_style=bold)
        current_row += 1
        summary.cell(row=current_row, column=1, value="Slots saved:").font = bold
        _write_cell(summary, current_row, 2, slots_saved, fill=blue_fill, font_style=Font(bold=True, color="006100"))
        current_row += 2

    # --- Flagged Rows Detail ---
    flagged = [a for a in assignments if a.get("is_new_ma") or a.get("is_new_com")]
    if flagged:
        summary.cell(row=current_row, column=1, value="Flagged Rows — Action Required").font = Font(bold=True, size=12, color="CC0000")
        current_row += 1
        _write_header_row(summary, current_row, ["Order Row", "Size", "Fabric", "Frame", "Embroidery", "Assigned MA", "Assigned COM", "Issue"])
        current_row += 1

        for a in flagged:
            issue = "New size — add MA to reference" if a.get("is_new_ma") else "New color combo — add COM to reference"
            _write_cell(summary, current_row, 1, a["row_num"], fill=yellow_fill, align_center=True)
            _write_cell(summary, current_row, 2, a.get("size", ""), fill=yellow_fill)
            _write_cell(summary, current_row, 3, a.get("fabric_colour", ""), fill=yellow_fill)
            _write_cell(summary, current_row, 4, a.get("frame_colour", ""), fill=yellow_fill)
            _write_cell(summary, current_row, 5, a.get("embroidery_colour", ""), fill=yellow_fill)
            _write_cell(summary, current_row, 6, a["assigned_ma"], fill=yellow_fill, font_style=bold)
            _write_cell(summary, current_row, 7, a["assigned_com"], fill=yellow_fill, font_style=bold)
            _write_cell(summary, current_row, 8, issue, fill=red_fill, font_style=bold)
            current_row += 1
    else:
        summary.cell(row=current_row, column=1, value="All rows matched reference — no action required").font = Font(bold=True, size=12, color="006100")
        summary.cell(row=current_row, column=1).fill = green_fill

    # --- Column widths for summary ---
    col_widths = [16, 12, 16, 20, 20, 14, 14, 35]
    for i, w in enumerate(col_widths, start=1):
        summary.column_dimensions[get_column_letter(i)].width = w

    # --- Save ---
    base, ext = os.path.splitext(original_path)
    output_path = f"{base}_assigned{ext}"
    wb.save(output_path)
    wb.close()
    return output_path
