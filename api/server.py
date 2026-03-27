"""
FastAPI backend for Micro Automation — Embroidery Stacker.
Handles Excel parsing, DST file uploads, and combo export.
Sessions are persisted in SQLite via database.py.
"""

# Load .env file if present (for local development)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

import asyncio
import json
import logging
import os
import secrets
import shutil
import uuid
import zipfile
from datetime import datetime, timezone
from io import BytesIO

import re

logger = logging.getLogger(__name__)

from fastapi import BackgroundTasks, FastAPI, File, Form, Response, UploadFile, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from starlette.middleware.gzip import GZipMiddleware

import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.core.excel_parser import (
    NameEntry, parse_excel, detect_columns, group_entries, generate_all_combos,
    detect_assign_columns, auto_assign_ma_com, export_assigned_excel,
)
from app.core.pipeline import export_all
from api.database import (
    get_db, create_session, get_session, update_session,
    delete_session as db_delete_session, list_sessions,
    get_session_dir, get_dst_dir, get_output_dir,
    cleanup_old_sessions,
    get_ma_reference, get_ma_lookup, upsert_ma_reference, clear_ma_reference,
    get_com_reference, get_com_lookup, get_max_com_per_ma,
    upsert_com_reference, clear_com_reference,
)
from api.auth import get_current_user

from contextlib import asynccontextmanager

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup: seed default MA/COM reference if empty
    from api.seed_data import seed_reference_if_empty
    seed_reference_if_empty()

    # Clean up old sessions from previous runs
    deleted = cleanup_old_sessions(max_age_hours=24)
    if deleted:
        print(f"[startup] Cleaned up {deleted} expired session(s)")
    yield

app = FastAPI(title="Micro Automation API", lifespan=lifespan)

# Semaphore to limit concurrent exports (prevents memory exhaustion)
_export_semaphore = asyncio.Semaphore(2)

# Compress JSON responses
app.add_middleware(GZipMiddleware, minimum_size=200)

# ---------------------------------------------------------------------------
# Security constants
# ---------------------------------------------------------------------------
MAX_EXCEL_SIZE = 10 * 1024 * 1024       # 10 MB
MAX_DST_SIZE = 5 * 1024 * 1024          # 5 MB per DST file
MAX_ZIP_SIZE = 100 * 1024 * 1024        # 100 MB compressed
MAX_ZIP_UNCOMPRESSED = 50 * 1024 * 1024  # 50 MB uncompressed (DST files are small)
MAX_ZIP_FILES = 500                      # max files in zip


def _safe_filename(name: str | None, fallback: str = "file") -> str:
    """Sanitise an uploaded filename to prevent path traversal."""
    if not name:
        return fallback
    # Take only the basename, strip any path separators
    name = os.path.basename(name)
    # Remove any remaining suspicious characters
    name = re.sub(r'[^\w.\-() ]', '_', name)
    if not name or name.startswith('.'):
        return fallback
    return name


_health_counter = 0

@app.get("/api/health")
async def health(background_tasks: BackgroundTasks):
    """Health check — also used to wake up Render free tier before user needs it."""
    global _health_counter
    _health_counter += 1
    if _health_counter % 100 == 0:
        background_tasks.add_task(cleanup_old_sessions, 24)
    return {"status": "ok"}


# CORS: allow localhost for dev + production frontend URLs from env
_cors_origins = ["http://localhost:3000", "http://localhost:5123"]
_frontend_url = os.environ.get("FRONTEND_URL")
if _frontend_url:
    # Support comma-separated URLs (e.g., "https://a.vercel.app,https://b.vercel.app")
    for url in _frontend_url.split(","):
        url = url.strip()
        if url:
            _cors_origins.append(url)

app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "DELETE", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization"],
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _require_session(session_id: str) -> dict:
    session = get_session(session_id)
    if session is None:
        raise HTTPException(404, "Session not found")
    return session


def _ensure_session(session_id: str = None, user_email: str = "local@dev") -> tuple:
    if session_id:
        session = get_session(session_id)
        if session:
            return session_id, session
    sid = session_id or secrets.token_urlsafe(16)
    session = create_session(sid, sid, user_email)
    return sid, session


def _entries_to_json(entries: list) -> list:
    return [
        {
            "program": e.program,
            "name_line1": e.name_line1,
            "name_line2": e.name_line2,
            "quantity": e.quantity,
            "com_no": e.com_no,
            "machine_program": e.machine_program,
        }
        for e in entries
    ]


def _entries_from_json(data: list) -> list:
    return [
        NameEntry(
            program=d["program"],
            name_line1=d["name_line1"],
            name_line2=d["name_line2"],
            quantity=d["quantity"],
            com_no=d["com_no"],
            machine_program=d["machine_program"],
        )
        for d in data
    ]


def _combos_to_json(combos: list) -> list:
    return [
        {
            "filename": c.filename,
            "machine_program": c.machine_program,
            "com_no": c.com_no,
            "part_number": c.part_number,
            "total_parts": c.total_parts,
            "slot_count": len(c.slots),
            "left_count": len(c.left_column),
            "right_count": len(c.right_column),
            "head_mode": c.head_mode or "",
            "slots": [
                {
                    "program": s.program,
                    "name_line1": s.name_line1,
                    "name_line2": s.name_line2,
                    "quantity": s.quantity,
                    "com_no": s.com_no,
                    "machine_program": s.machine_program,
                }
                for s in c.slots
            ],
        }
        for c in combos
    ]


def _build_groups_response(entries, combos, groups) -> list:
    # Pre-build lookup to avoid O(n*m) scan
    from collections import defaultdict as _defaultdict
    combo_by_group = _defaultdict(list)
    for c in combos:
        combo_by_group[(c.machine_program, c.com_no)].append(c)

    groups_data = []
    for g in groups:
        g_combos = combo_by_group.get((g.machine_program, g.com_no), [])
        groups_data.append({
            "machine_program": g.machine_program,
            "com_no": g.com_no,
            "entry_count": len(g.entries),
            "total_slots": g.total_slots,
            "combos": [
                {
                    "filename": c.filename,
                    "part_number": c.part_number,
                    "total_parts": c.total_parts,
                    "slot_count": len(c.slots),
                    "left_count": len(c.left_column),
                    "right_count": len(c.right_column),
                    "head_mode": c.head_mode or "",
                    "slots": [
                        {"program": s.program, "name_line1": s.name_line1,
                         "name_line2": s.name_line2, "quantity": s.quantity}
                        for s in c.slots
                    ],
                }
                for c in g_combos
            ],
        })
    return groups_data


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@app.post("/api/detect-columns")
async def detect_columns_endpoint(
    file: UploadFile = File(...),
    session_id: str = Form(None),
    user: dict = Depends(get_current_user),
):
    """Upload Excel and auto-detect column mapping. Returns mapping + preview for confirmation."""
    sid, session = _ensure_session(session_id, user["email"])

    session_dir = get_session_dir(sid)
    content = await file.read()
    if len(content) > MAX_EXCEL_SIZE:
        raise HTTPException(413, f"Excel file too large (max {MAX_EXCEL_SIZE // 1024 // 1024}MB)")
    safe_name = _safe_filename(file.filename, "order.xlsx")
    excel_path = os.path.join(session_dir, safe_name)
    with open(excel_path, "wb") as f:
        f.write(content)

    update_session(sid, excel_filename=file.filename)

    detection = detect_columns(excel_path)

    return {
        "session_id": sid,
        "excel_filename": file.filename,
        "headers": detection.headers,
        "preview_rows": detection.preview_rows,
        "detected_mapping": detection.detected_mapping,
        "confidence": detection.confidence,
    }


@app.post("/api/parse-excel")
async def parse_excel_endpoint(
    file: UploadFile = File(None),
    session_id: str = Form(None),
    column_map: str = Form(None),
    optimize_heads: bool = Form(False),
    user: dict = Depends(get_current_user),
):
    sid, session = _ensure_session(session_id, user["email"])

    # Parse column_map if provided
    cmap = None
    if column_map:
        try:
            cmap = json.loads(column_map)
            # Ensure all values are ints
            cmap = {k: int(v) for k, v in cmap.items()}
        except (json.JSONDecodeError, ValueError):
            raise HTTPException(400, "Invalid column_map JSON")

    # If a new file is uploaded, save it. Otherwise use existing Excel in session.
    session_dir = get_session_dir(sid)
    if file:
        content = await file.read()
        if len(content) > MAX_EXCEL_SIZE:
            raise HTTPException(413, f"Excel file too large (max {MAX_EXCEL_SIZE // 1024 // 1024}MB)")
        safe_name = _safe_filename(file.filename, "order.xlsx")
        excel_path = os.path.join(session_dir, safe_name)
        with open(excel_path, "wb") as f:
            f.write(content)
        update_session(sid, excel_filename=file.filename)
    else:
        # Find existing Excel file in session dir
        excel_files = [f for f in os.listdir(session_dir) if f.endswith(('.xlsx', '.xls'))]
        if not excel_files:
            raise HTTPException(400, "No Excel file found. Upload a file or use detect-columns first.")
        excel_path = os.path.join(session_dir, excel_files[0])

    result = parse_excel(excel_path, column_map=cmap)
    if not result.entries:
        update_session(sid, entries_json=[], groups_json=[], combos_json=[])
        return {"session_id": sid, "entries_count": 0, "groups": [], "combos": [], "warnings": result.warnings}

    groups = group_entries(result.entries)
    combos = generate_all_combos(result.entries, optimize_heads=optimize_heads)

    entries_data = _entries_to_json(result.entries)
    groups_data = _build_groups_response(result.entries, combos, groups)
    combos_data = _combos_to_json(combos)

    # Calculate head optimization stats
    even_file_count = sum(1 for c in combos if c.head_mode == "2-HEAD")
    odd_file_count = sum(1 for c in combos if c.head_mode == "1-HEAD")
    total_original_slots = sum(e.quantity for e in result.entries)
    total_optimized_slots = sum(len(c.slots) for c in combos)
    slots_saved = total_original_slots - total_optimized_slots if optimize_heads else 0

    update_session(
        sid,
        entries_json=entries_data,
        groups_json=groups_data,
        combos_json=combos_data,
        optimize_heads=1 if optimize_heads else 0,
    )

    return {
        "session_id": sid,
        "entries_count": len(result.entries),
        "total_slots": total_original_slots,
        "optimized_slots": total_optimized_slots,
        "slots_saved": slots_saved,
        "even_file_count": even_file_count,
        "odd_file_count": odd_file_count,
        "groups": groups_data,
        "combo_count": len(combos),
        "warnings": result.warnings,
        "entries_preview": entries_data,
    }


@app.post("/api/upload-dst")
async def upload_dst_endpoint(
    session_id: str = Form(...),
    files: list[UploadFile] = File(None),
    zip_file: UploadFile = File(None),
    user: dict = Depends(get_current_user),
):
    _require_session(session_id)
    dst_dir = get_dst_dir(session_id)
    found_programs = []

    if zip_file and zip_file.filename and zip_file.filename.lower().endswith(".zip"):
        content = await zip_file.read()
        if len(content) > MAX_ZIP_SIZE:
            logger.warning("ZIP upload rejected: %d bytes exceeds %d limit", len(content), MAX_ZIP_SIZE)
            raise HTTPException(413, f"ZIP file too large (max {MAX_ZIP_SIZE // 1024 // 1024}MB)")
        with zipfile.ZipFile(BytesIO(content)) as zf:
            # Zip bomb protection: check uncompressed size and file count
            infos = zf.infolist()
            if len(infos) > MAX_ZIP_FILES:
                raise HTTPException(400, f"ZIP contains too many files (max {MAX_ZIP_FILES})")
            total_uncompressed = sum(i.file_size for i in infos)
            if total_uncompressed > MAX_ZIP_UNCOMPRESSED:
                raise HTTPException(413, f"ZIP uncompressed size too large (max {MAX_ZIP_UNCOMPRESSED // 1024 // 1024}MB)")
            for name in zf.namelist():
                # Reject nested ZIPs
                if name.lower().endswith(".zip"):
                    raise HTTPException(400, "Nested ZIP files are not allowed")
                basename = os.path.basename(name)
                if basename.lower().endswith(".dst") and not basename.startswith("."):
                    safe_name = _safe_filename(basename, "unnamed.dst")
                    target = os.path.join(dst_dir, safe_name)
                    # Verify path stays within dst_dir
                    if not os.path.realpath(target).startswith(os.path.realpath(dst_dir)):
                        continue
                    with open(target, "wb") as out:
                        out.write(zf.read(name))

    if files:
        for f in files:
            if f.filename and f.filename.lower().endswith(".dst"):
                content = await f.read()
                if len(content) > MAX_DST_SIZE:
                    continue  # skip oversized individual DST files
                safe_name = _safe_filename(f.filename, "unnamed.dst")
                target = os.path.join(dst_dir, safe_name)
                if not os.path.realpath(target).startswith(os.path.realpath(dst_dir)):
                    continue
                with open(target, "wb") as out:
                    out.write(content)

    for fname in os.listdir(dst_dir):
        if fname.lower().endswith(".dst"):
            stem = os.path.splitext(fname)[0]
            try:
                found_programs.append(int(stem))
            except ValueError:
                pass

    found_programs.sort()
    session = update_session(session_id, dst_programs_json=found_programs)

    entries_data = session.get("entries_json") or []
    needed = set(e["program"] for e in entries_data)
    found_set = set(found_programs)
    missing = sorted(needed - found_set)

    return {
        "session_id": session_id,
        "uploaded_count": len(found_programs),
        "found_programs": found_programs,
        "needed_count": len(needed),
        "missing_programs": missing,
        "all_matched": len(missing) == 0,
    }


@app.post("/api/export")
async def export_endpoint(
    session_id: str = Form(...),
    selected_filenames: str = Form(""),
    gap_mm: float = Form(3.0),
    column_gap_mm: float = Form(5.0),
    user: dict = Depends(get_current_user),
):
    session = _require_session(session_id)
    entries_data = session.get("entries_json") or []
    if not entries_data:
        raise HTTPException(400, "No combos parsed yet")

    entries = _entries_from_json(entries_data)
    optimize_heads = bool(session.get("optimize_heads", 0))
    combos = generate_all_combos(entries, optimize_heads=optimize_heads)

    selected_set = set(f.strip() for f in selected_filenames.split(",") if f.strip())
    if selected_set:
        combos_to_export = [c for c in combos if c.filename in selected_set]
    else:
        combos_to_export = combos

    if not combos_to_export:
        raise HTTPException(400, "No combos selected")

    import tempfile

    output_dir = get_output_dir(session_id)
    dst_dir = get_dst_dir(session_id)

    # Limit concurrent exports to prevent memory exhaustion
    try:
        await asyncio.wait_for(_export_semaphore.acquire(), timeout=60)
    except asyncio.TimeoutError:
        raise HTTPException(503, "Server is busy with other exports. Please try again shortly.")

    # Export to temp dir first — only replace output_dir on success
    # Run CPU-bound combining in a thread so we don't block the event loop
    tmp_dir = tempfile.mkdtemp(prefix="combo_export_")
    try:
        results = await asyncio.to_thread(
            export_all,
            combos_to_export, dst_dir, tmp_dir,
            gap_mm=gap_mm, column_gap_mm=column_gap_mm, overwrite=True,
        )

        success = [r for r in results if r.success]
        failed = [r for r in results if not r.success]

        if not success:
            raise HTTPException(500, f"All exports failed. First error: {failed[0].error if failed else 'unknown'}")

        # Success — move files from temp to output dir
        for f in os.listdir(output_dir):
            os.remove(os.path.join(output_dir, f))
        for r in success:
            dest = os.path.join(output_dir, os.path.basename(r.output_path))
            shutil.move(r.output_path, dest)
            r.output_path = dest
    finally:
        _export_semaphore.release()
        shutil.rmtree(tmp_dir, ignore_errors=True)

    update_session(
        session_id,
        exported=1,
        exported_at=datetime.now(timezone.utc).isoformat(),
        gap_mm=gap_mm,
        column_gap_mm=column_gap_mm,
    )

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for r in success:
            zf.write(r.output_path, os.path.basename(r.output_path))
    zip_buffer.seek(0)

    session_name = re.sub(r'[^\w\-. ]', '_', session.get("name", session_id))
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="combos_{session_name}.zip"',
            "X-Export-Success": str(len(success)),
            "X-Export-Failed": str(len(failed)),
        },
    )


# ---------------------------------------------------------------------------
# MA Reference table
# ---------------------------------------------------------------------------

@app.get("/api/ma-reference")
async def get_ma_reference_endpoint(response: Response, user: dict = Depends(get_current_user)):
    """Return the stored MA reference lookup table."""
    response.headers["Cache-Control"] = "private, max-age=300"
    mappings = get_ma_reference()
    return {"count": len(mappings), "mappings": mappings}


@app.post("/api/ma-reference/upload")
async def upload_ma_reference_endpoint(
    file: UploadFile = File(...),
    mode: str = Form("replace"),
    user: dict = Depends(get_current_user),
):
    """Upload an Excel file containing size → MA number mappings.

    Expected format: Column A = size, Column F = MA number.
    mode='replace' clears existing data first; mode='append' only adds new entries.
    """
    from openpyxl import load_workbook
    import tempfile

    content = await file.read()
    if len(content) > MAX_EXCEL_SIZE:
        raise HTTPException(413, f"File too large (max {MAX_EXCEL_SIZE // 1024 // 1024}MB)")

    # Write to temp file for openpyxl
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        wb = load_workbook(tmp_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    finally:
        os.remove(tmp_path)

    if len(rows) < 2:
        raise HTTPException(400, "Excel file has no data rows")

    # Parse columns:
    #   A (0) = size, B (1) = fabric, C (2) = embroidery, D (3) = frame
    #   F (5) = MA number, G (6) = COM number
    mappings = []
    com_mappings = []
    seen_sizes = {}  # normalized → mapping dict
    warnings = []

    for row_num, row in enumerate(rows[1:], start=2):
        if len(row) < 6:
            continue
        size_raw = str(row[0] or "").strip()
        ma_raw = str(row[5] or "").strip()
        if not size_raw or not ma_raw:
            continue

        size_normalized = re.sub(r'\s+', '', size_raw).lower()

        if size_normalized not in seen_sizes:
            seen_sizes[size_normalized] = {
                "size_normalized": size_normalized,
                "size_display": size_raw,
                "ma_number": ma_raw,
            }
        elif seen_sizes[size_normalized]["ma_number"] != ma_raw:
            warnings.append(
                f"Row {row_num}: size '{size_raw}' mapped to '{ma_raw}' "
                f"but already mapped to '{seen_sizes[size_normalized]['ma_number']}' — keeping first"
            )

        # Parse COM + color columns (B=fabric, C=name/embroidery, D=frame, G=com)
        # Reference Excel: C=embroidery, D=frame
        # Order Excel:     J=frame,      K=embroidery
        # We store as frame_colour / embroidery_colour to match auto-assign lookup keys
        fabric_raw = str(row[1] or "").strip().title() if len(row) > 1 else ""
        embroidery_raw = str(row[2] or "").strip().title() if len(row) > 2 else ""
        frame_raw = str(row[3] or "").strip().title() if len(row) > 3 else ""
        com_raw = row[6] if len(row) > 6 else None

        if fabric_raw and embroidery_raw and frame_raw and com_raw is not None:
            try:
                com_num = int(com_raw)
                com_mappings.append({
                    "ma_number": ma_raw,
                    "com_number": com_num,
                    "fabric_colour": fabric_raw,
                    "embroidery_colour": embroidery_raw,
                    "frame_colour": frame_raw,
                })
            except (ValueError, TypeError):
                warnings.append(f"Row {row_num}: invalid COM number '{com_raw}', skipped COM mapping")

    mappings = list(seen_sizes.values())
    if not mappings:
        raise HTTPException(400, "No valid size → MA mappings found in the file")

    # Replace clears everything first; append only adds new entries (upsert handles duplicates)
    if mode == "replace":
        clear_ma_reference()
        clear_com_reference()
    count = upsert_ma_reference(mappings)
    com_count = upsert_com_reference(com_mappings) if com_mappings else 0

    return {
        "count": count,
        "com_count": com_count,
        "mappings": mappings,
        "warnings": warnings,
    }


@app.delete("/api/ma-reference")
async def delete_ma_reference_endpoint(user: dict = Depends(get_current_user)):
    """Clear all MA and COM reference data."""
    deleted_ma = clear_ma_reference()
    deleted_com = clear_com_reference()
    return {"deleted_ma": deleted_ma, "deleted_com": deleted_com}


@app.get("/api/com-reference")
async def get_com_reference_endpoint(response: Response, user: dict = Depends(get_current_user)):
    """Return all stored COM reference entries."""
    response.headers["Cache-Control"] = "private, max-age=300"
    from api.database import get_com_reference as _get_com_ref
    entries = _get_com_ref()
    return {"count": len(entries), "entries": entries}


@app.post("/api/ma-reference/add")
async def add_single_ma_endpoint(
    size_normalized: str = Form(...),
    size_display: str = Form(...),
    ma_number: str = Form(...),
    user: dict = Depends(get_current_user),
):
    """Add or update a single MA reference entry."""
    from api.database import add_single_ma
    entry = add_single_ma(size_normalized, size_display, ma_number)
    return {"ok": True, "entry": entry}


@app.put("/api/ma-reference/{entry_id}")
async def update_ma_reference_endpoint(
    entry_id: int,
    size_normalized: str = Form(None),
    size_display: str = Form(None),
    ma_number: str = Form(None),
    user: dict = Depends(get_current_user),
):
    """Update a single MA reference entry by ID."""
    from api.database import update_ma_reference_entry
    kwargs = {}
    if size_normalized is not None:
        kwargs["size_normalized"] = size_normalized
    if size_display is not None:
        kwargs["size_display"] = size_display
    if ma_number is not None:
        kwargs["ma_number"] = ma_number
    entry = update_ma_reference_entry(entry_id, **kwargs)
    if entry is None:
        raise HTTPException(404, "MA reference entry not found")
    return {"ok": True, "entry": entry}


@app.delete("/api/ma-reference/{entry_id}")
async def delete_single_ma_endpoint(
    entry_id: int,
    user: dict = Depends(get_current_user),
):
    """Delete a single MA reference entry by ID."""
    from api.database import delete_ma_reference_entry
    deleted = delete_ma_reference_entry(entry_id)
    if not deleted:
        raise HTTPException(404, "MA reference entry not found")
    return {"ok": True}


@app.post("/api/com-reference/add")
async def add_single_com_endpoint(
    ma_number: str = Form(...),
    com_number: int = Form(...),
    fabric_colour: str = Form(...),
    embroidery_colour: str = Form(...),
    frame_colour: str = Form(...),
    user: dict = Depends(get_current_user),
):
    """Add or update a single COM reference entry."""
    from api.database import add_single_com
    entry = add_single_com(ma_number, com_number, fabric_colour, embroidery_colour, frame_colour)
    return {"ok": True, "entry": entry}


@app.put("/api/com-reference/{entry_id}")
async def update_com_reference_endpoint(
    entry_id: int,
    ma_number: str = Form(None),
    com_number: int = Form(None),
    fabric_colour: str = Form(None),
    embroidery_colour: str = Form(None),
    frame_colour: str = Form(None),
    user: dict = Depends(get_current_user),
):
    """Update a single COM reference entry by ID."""
    from api.database import update_com_reference_entry
    kwargs = {}
    if ma_number is not None:
        kwargs["ma_number"] = ma_number
    if com_number is not None:
        kwargs["com_number"] = com_number
    if fabric_colour is not None:
        kwargs["fabric_colour"] = fabric_colour
    if embroidery_colour is not None:
        kwargs["embroidery_colour"] = embroidery_colour
    if frame_colour is not None:
        kwargs["frame_colour"] = frame_colour
    entry = update_com_reference_entry(entry_id, **kwargs)
    if entry is None:
        raise HTTPException(404, "COM reference entry not found")
    return {"ok": True, "entry": entry}


@app.delete("/api/com-reference/{entry_id}")
async def delete_single_com_endpoint(
    entry_id: int,
    user: dict = Depends(get_current_user),
):
    """Delete a single COM reference entry by ID."""
    from api.database import delete_com_reference_entry
    deleted = delete_com_reference_entry(entry_id)
    if not deleted:
        raise HTTPException(404, "COM reference entry not found")
    return {"ok": True}


# ---------------------------------------------------------------------------
# Auto-assign MA & COM
# ---------------------------------------------------------------------------

@app.post("/api/detect-assign-columns")
async def detect_assign_columns_endpoint(
    file: UploadFile = File(...),
    session_id: str = Form(None),
    user: dict = Depends(get_current_user),
):
    """Upload Excel and auto-detect columns for MA/COM assignment (size, fabric, frame, embroidery)."""
    sid, session = _ensure_session(session_id, user["email"])
    session_dir = get_session_dir(sid)

    content = await file.read()
    if len(content) > MAX_EXCEL_SIZE:
        raise HTTPException(413, f"Excel file too large (max {MAX_EXCEL_SIZE // 1024 // 1024}MB)")
    safe_name = _safe_filename(file.filename, "order.xlsx")
    excel_path = os.path.join(session_dir, safe_name)
    with open(excel_path, "wb") as f:
        f.write(content)

    update_session(sid, excel_filename=file.filename)

    detection = detect_assign_columns(excel_path)

    return {
        "session_id": sid,
        "excel_filename": file.filename,
        "headers": detection.headers,
        "preview_rows": detection.preview_rows,
        "detected_mapping": detection.detected_mapping,
        "confidence": detection.confidence,
    }


@app.post("/api/auto-assign")
async def auto_assign_endpoint(
    session_id: str = Form(...),
    column_map: str = Form(None),
    user: dict = Depends(get_current_user),
):
    """Run auto-assignment of MA & COM using confirmed column mapping."""
    session = _require_session(session_id)
    session_dir = get_session_dir(session_id)

    # Find existing Excel file in session dir
    excel_files = [f for f in os.listdir(session_dir) if f.endswith(('.xlsx', '.xls'))]
    if not excel_files:
        raise HTTPException(400, "No Excel file found. Upload a file first.")
    excel_path = os.path.join(session_dir, excel_files[0])

    cmap = None
    if column_map:
        try:
            cmap = json.loads(column_map)
            cmap = {k: int(v) for k, v in cmap.items()}
        except (json.JSONDecodeError, ValueError):
            raise HTTPException(400, "Invalid column_map JSON")

    # Load MA + COM reference lookups if available
    ma_ref = get_ma_lookup()
    com_ref = get_com_lookup()
    max_coms = get_max_com_per_ma()
    result = auto_assign_ma_com(
        excel_path, column_map=cmap,
        ma_lookup=ma_ref or None,
        com_lookup=com_ref or None,
        max_com_per_ma=max_coms or None,
    )

    # Store assignments in session for later use (download / apply)
    update_session(
        session_id,
        assign_result_json={
            "assignments": result.assignments,
            "ma_summary": result.ma_summary,
            "com_summary": result.com_summary,
            "column_map": cmap or result.detected_mapping,
        },
    )

    return {
        "session_id": session_id,
        "assignments_count": len(result.assignments),
        "ma_summary": result.ma_summary,
        "com_summary": result.com_summary,
        "warnings": result.warnings,
        "assignments_preview": result.assignments[:20],
    }


@app.post("/api/download-assigned-excel")
async def download_assigned_excel_endpoint(
    session_id: str = Form(...),
    user: dict = Depends(get_current_user),
):
    """Download an Excel file with MA & COM columns filled in."""
    session = _require_session(session_id)
    session_dir = get_session_dir(session_id)

    assign_data = session.get("assign_result_json")
    if not assign_data or not assign_data.get("assignments"):
        raise HTTPException(400, "No auto-assign results found. Run auto-assign first.")

    excel_files = [f for f in os.listdir(session_dir) if f.endswith(('.xlsx', '.xls'))]
    if not excel_files:
        raise HTTPException(400, "No Excel file found.")
    excel_path = os.path.join(session_dir, excel_files[0])

    output_path = export_assigned_excel(
        excel_path,
        assign_data["assignments"],
        ma_summary=assign_data.get("ma_summary"),
        com_summary=assign_data.get("com_summary"),
    )

    with open(output_path, "rb") as f:
        file_bytes = f.read()
    os.remove(output_path)

    original_name = session.get("excel_filename", "order.xlsx")
    base, ext = os.path.splitext(original_name)
    download_name = f"{base}_with_MA_COM{ext}"

    return StreamingResponse(
        BytesIO(file_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
    )


@app.post("/api/apply-assignments")
async def apply_assignments_endpoint(
    session_id: str = Form(...),
    column_map: str = Form(None),
    optimize_heads: bool = Form(False),
    user: dict = Depends(get_current_user),
):
    """Apply auto-assigned MA/COM and parse into combos (combines auto-assign + parse-excel)."""
    session = _require_session(session_id)
    session_dir = get_session_dir(session_id)

    assign_data = session.get("assign_result_json")
    if not assign_data or not assign_data.get("assignments"):
        raise HTTPException(400, "No auto-assign results found. Run auto-assign first.")

    excel_files = [f for f in os.listdir(session_dir) if f.endswith(('.xlsx', '.xls'))]
    if not excel_files:
        raise HTTPException(400, "No Excel file found.")
    excel_path = os.path.join(session_dir, excel_files[0])

    # Write assigned Excel, then parse it using the standard column map
    assigned_path = export_assigned_excel(
        excel_path,
        assign_data["assignments"],
        ma_summary=assign_data.get("ma_summary"),
        com_summary=assign_data.get("com_summary"),
        optimize_heads=optimize_heads,
    )

    # Parse column_map for the standard 6-field mapping if provided
    cmap = None
    if column_map:
        try:
            cmap = json.loads(column_map)
            cmap = {k: int(v) for k, v in cmap.items()}
        except (json.JSONDecodeError, ValueError):
            raise HTTPException(400, "Invalid column_map JSON")

    # Always auto-detect columns on the assigned Excel (MA/COM columns were appended)
    if cmap is None:
        detected = detect_columns(assigned_path)
        cmap = detected.detected_mapping

    result = parse_excel(assigned_path, column_map=cmap)
    os.remove(assigned_path)

    if not result.entries:
        update_session(session_id, entries_json=[], groups_json=[], combos_json=[])
        return {"session_id": session_id, "entries_count": 0, "groups": [], "combos": [], "warnings": result.warnings}

    groups = group_entries(result.entries)
    combos = generate_all_combos(result.entries, optimize_heads=optimize_heads)

    entries_data = _entries_to_json(result.entries)
    groups_data = _build_groups_response(result.entries, combos, groups)
    combos_data = _combos_to_json(combos)

    even_file_count = sum(1 for c in combos if c.head_mode == "2-HEAD")
    odd_file_count = sum(1 for c in combos if c.head_mode == "1-HEAD")
    total_original_slots = sum(e.quantity for e in result.entries)
    total_optimized_slots = sum(len(c.slots) for c in combos)
    slots_saved = total_original_slots - total_optimized_slots if optimize_heads else 0

    update_session(
        session_id,
        entries_json=entries_data,
        groups_json=groups_data,
        combos_json=combos_data,
        optimize_heads=1 if optimize_heads else 0,
    )

    return {
        "session_id": session_id,
        "entries_count": len(result.entries),
        "total_slots": total_original_slots,
        "optimized_slots": total_optimized_slots,
        "slots_saved": slots_saved,
        "even_file_count": even_file_count,
        "odd_file_count": odd_file_count,
        "groups": groups_data,
        "combo_count": len(combos),
        "warnings": result.warnings,
        "entries_preview": entries_data,
    }


# ---------------------------------------------------------------------------
# Session management
# ---------------------------------------------------------------------------

@app.get("/api/session/{session_id}/status")
async def session_status(session_id: str, user: dict = Depends(get_current_user)):
    session = _require_session(session_id)
    entries = session.get("entries_json") or []
    combos = session.get("combos_json") or []
    dst = session.get("dst_programs_json") or []
    return {
        "session_id": session_id,
        "has_excel": len(entries) > 0,
        "entries_count": len(entries),
        "combo_count": len(combos),
        "dst_count": len(dst),
        "exported": session.get("exported", False),
    }


@app.get("/api/session/{session_id}/full")
async def session_full(session_id: str, user: dict = Depends(get_current_user)):
    """Load complete session state for switching between sessions."""
    session = _require_session(session_id)
    entries = session.get("entries_json") or []
    groups = session.get("groups_json") or []
    dst_programs = session.get("dst_programs_json") or []

    needed = set(e["program"] for e in entries)
    found_set = set(dst_programs)
    missing = sorted(needed - found_set)

    return {
        "session_id": session_id,
        "name": session["name"],
        "entries_count": len(entries),
        "total_slots": sum(e.get("quantity", 1) for e in entries),
        "groups": groups,
        "combo_count": len(session.get("combos_json") or []),
        "entries_preview": entries,
        "dst_count": len(dst_programs),
        "dst_all_matched": len(missing) == 0,
        "missing_programs": missing,
        "exported": session.get("exported", False),
        "exported_at": session.get("exported_at"),
        "excel_filename": session.get("excel_filename"),
        "gap_mm": session.get("gap_mm", 3.0),
        "column_gap_mm": session.get("column_gap_mm", 5.0),
        "warnings": [],
    }


@app.get("/api/sessions")
async def list_sessions_endpoint(user: dict = Depends(get_current_user)):
    sessions = list_sessions(user["email"])
    for s in sessions:
        s["session_id"] = s.pop("id")
    return sessions


@app.post("/api/session/name")
async def set_session_name(
    session_id: str = Form(...), name: str = Form(...),
    user: dict = Depends(get_current_user),
):
    _require_session(session_id)
    update_session(session_id, name=name)
    return {"ok": True}


@app.delete("/api/session/{session_id}")
async def delete_session_endpoint(session_id: str, user: dict = Depends(get_current_user)):
    _require_session(session_id)
    session_dir = get_session_dir(session_id)
    if os.path.isdir(session_dir):
        shutil.rmtree(session_dir, ignore_errors=True)
    db_delete_session(session_id)
    return {"ok": True}


@app.delete("/api/session/{session_id}/excel")
async def remove_excel(session_id: str, user: dict = Depends(get_current_user)):
    """Remove Excel data, resetting parsed entries/combos/groups."""
    _require_session(session_id)
    update_session(
        session_id,
        excel_filename=None,
        entries_json=[],
        groups_json=[],
        combos_json=[],
        exported=0,
        exported_at=None,
    )
    output_dir = get_output_dir(session_id)
    if os.path.isdir(output_dir):
        for f in os.listdir(output_dir):
            os.remove(os.path.join(output_dir, f))
    return {"ok": True}


# ---------------------------------------------------------------------------
# Dev endpoint
# ---------------------------------------------------------------------------

@app.get("/api/dev/load-sample")
async def dev_load_sample(user: dict = Depends(get_current_user)):
    """Dev endpoint: auto-load real Excel + DST files for testing."""
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    test_data = os.path.join(project_root, "test_data")
    excel_path = os.path.join(test_data, "nameorder_04032026-2 add column (1).xlsx")
    if not os.path.isfile(excel_path):
        excel_path = os.path.expanduser("~/Downloads/nameorder_04032026-2 add column (1).xlsx")
    if not os.path.isfile(excel_path):
        excel_path = os.path.expanduser("~/Downloads/nameorder_04032026-2 add column.xlsx")
    dst_zip = os.path.join(test_data, "programs_Micro.zip")
    if not os.path.isfile(dst_zip):
        dst_zip = os.path.expanduser("~/Downloads/programs_Micro.zip")

    if not os.path.isfile(excel_path):
        raise HTTPException(404, f"Excel not found: {excel_path}")

    sid, session = _ensure_session("dev")

    result = parse_excel(excel_path)
    groups = group_entries(result.entries)
    combos = generate_all_combos(result.entries)

    entries_data = _entries_to_json(result.entries)
    groups_data = _build_groups_response(result.entries, combos, groups)
    combos_data = _combos_to_json(combos)

    update_session(
        sid,
        name="Dev Sample",
        excel_filename=os.path.basename(excel_path),
        entries_json=entries_data,
        groups_json=groups_data,
        combos_json=combos_data,
    )

    dst_dir = get_dst_dir(sid)
    if os.path.isfile(dst_zip):
        import zipfile as zf_mod
        with zf_mod.ZipFile(dst_zip) as z:
            for name in z.namelist():
                basename = os.path.basename(name)
                if basename.lower().endswith(".dst") and not basename.startswith("."):
                    with open(os.path.join(dst_dir, basename), "wb") as out:
                        out.write(z.read(name))

    dst_folder = os.path.join(project_root, "test_real_dst")
    if os.path.isdir(dst_folder):
        for fname in os.listdir(dst_folder):
            if fname.lower().endswith(".dst"):
                src = os.path.join(dst_folder, fname)
                tgt = os.path.join(dst_dir, fname)
                if not os.path.exists(tgt):
                    shutil.copy2(src, tgt)

    found = []
    for fname in os.listdir(dst_dir):
        if fname.lower().endswith(".dst"):
            try:
                found.append(int(os.path.splitext(fname)[0]))
            except ValueError:
                pass
    found.sort()
    update_session(sid, dst_programs_json=found)

    return {
        "session_id": sid,
        "entries_count": len(result.entries),
        "total_slots": sum(e.quantity for e in result.entries),
        "groups": groups_data,
        "combo_count": len(combos),
        "warnings": result.warnings,
        "entries_preview": entries_data,
        "dst_count": len(found),
        "dst_all_matched": True,
    }


# ---------------------------------------------------------------------------
# Version endpoint (for update notifications)
# ---------------------------------------------------------------------------

APP_VERSION = "1.0.0"
GITHUB_REPO = "shaanpawa/embroidery-combiner"
_version_cache: dict = {}
_version_cache_time: float = 0


@app.get("/api/version")
async def version_endpoint():
    """Return current app version and latest available version from GitHub."""
    import time
    global _version_cache, _version_cache_time

    result = {"version": APP_VERSION, "latest": None, "update_url": None, "update_available": False, "installer_url": None}

    # Cache GitHub check for 1 hour
    if _version_cache and time.time() - _version_cache_time < 3600:
        result.update(_version_cache)
        return result

    # Non-blocking check for latest release
    try:
        import urllib.request
        url = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
        req = urllib.request.Request(url, headers={"Accept": "application/vnd.github.v3+json"})
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read())
            latest = data.get("tag_name", "").lstrip("v")
            html_url = data.get("html_url", "")
            # Find the installer asset (.exe) for desktop auto-update
            # Prefer MicroAutomation_Setup over legacy EmbroideryC.exe
            installer_url = None
            for asset in data.get("assets", []):
                name = asset.get("name", "")
                if name.startswith("MicroAutomation_Setup") and name.endswith(".exe"):
                    installer_url = asset.get("browser_download_url")
                    break
                elif name.endswith(".exe") and not installer_url:
                    installer_url = asset.get("browser_download_url")
            _version_cache = {
                "latest": latest,
                "update_url": html_url,
                "installer_url": installer_url,
                "update_available": latest != APP_VERSION and latest > APP_VERSION,
            }
            _version_cache_time = time.time()
            result.update(_version_cache)
    except Exception:
        pass  # No internet or repo not set up yet — gracefully skip

    return result


@app.get("/api/update/download")
async def download_update():
    """Download the latest installer from GitHub to a temp directory (desktop mode only)."""
    import tempfile
    import urllib.request

    if os.environ.get("DESKTOP_MODE", "").lower() != "true":
        raise HTTPException(status_code=404, detail="Only available in desktop mode")

    # Get the installer URL from version check
    version_info = await version_endpoint()
    if not version_info.get("update_available") or not version_info.get("installer_url"):
        raise HTTPException(status_code=404, detail="No update available")

    installer_url = version_info["installer_url"]
    latest = version_info["latest"]

    try:
        temp_dir = os.path.join(tempfile.gettempdir(), "MicroAutomation_Update")
        os.makedirs(temp_dir, exist_ok=True)
        installer_path = os.path.join(temp_dir, f"MicroAutomation_Setup_v{latest}.exe")

        # Download if not already cached
        if not os.path.exists(installer_path):
            req = urllib.request.Request(installer_url)
            with urllib.request.urlopen(req, timeout=120) as resp:
                with open(installer_path, "wb") as f:
                    f.write(resp.read())

        return {"path": installer_path, "version": latest}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Download failed: {e}")


@app.post("/api/update/install")
async def install_update():
    """Launch the downloaded installer and shut down the current app (desktop mode only)."""
    import subprocess
    import tempfile
    import glob as glob_mod

    if os.environ.get("DESKTOP_MODE", "").lower() != "true":
        raise HTTPException(status_code=404, detail="Only available in desktop mode")

    temp_dir = os.path.join(tempfile.gettempdir(), "MicroAutomation_Update")
    installers = sorted(glob_mod.glob(os.path.join(temp_dir, "MicroAutomation_Setup_v*.exe")))
    if not installers:
        raise HTTPException(status_code=404, detail="No downloaded installer found")

    installer_path = installers[-1]  # Latest version

    try:
        # Launch installer in silent mode and exit
        subprocess.Popen([installer_path, "/SILENT"], creationflags=getattr(subprocess, "DETACHED_PROCESS", 0))
        # Schedule shutdown
        import threading
        threading.Timer(1.0, lambda: os._exit(0)).start()
        return {"status": "installing", "path": installer_path}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Install launch failed: {e}")


# ---------------------------------------------------------------------------
# Static file serving for desktop/local mode
# ---------------------------------------------------------------------------

if os.environ.get("DESKTOP_MODE", "").lower() == "true":
    from fastapi.staticfiles import StaticFiles

    _app_root = os.environ.get("MICRO_APP_ROOT", os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    _static_dir = os.path.join(_app_root, "static_web")
    if os.path.isdir(_static_dir):
        # Must be mounted LAST so API routes take priority
        app.mount("/", StaticFiles(directory=_static_dir, html=True), name="frontend")
